"""
Adapters and interfaces for the aggregator to a spreadsheet
"""
import logging
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .processor import DataIterator

logger = logging.getLogger(__name__)  # pylint: disable=invalid-name


class OpenpyxlDataSource(DataIterator):
    """ Provide a datasource based on openpyxl to read spreadsheets """

    def __init__(self, filename: Path, header_row: int, data_row: int = None):
        self.header_row = header_row
        if not data_row:
            self.data_row = header_row + 1
        else:
            self.data_row = data_row

        # Open workbook to worksheet
        source_wb = load_workbook(filename)
        source_ws = source_wb.active

        # Map columns to names
        row = next(source_ws.iter_rows(min_row=header_row))
        self.mapping = {}
        for cell in row:
            if cell.value not in self.mapping.keys():
                self.mapping[cell.column] = cell.value
            else:
                raise ValueError(
                    "Column '{}' already defined in cell {}, rename column in {}".format(
                        cell.value,
                        get_column_letter(self.mapping[cell.value]) + str(header_row),
                        cell.coordinate,
                    )
                )

        logger.debug("Mapping: %s", self.mapping)

        # Store row generator:
        self.rows = source_ws.iter_rows(min_row=self.data_row)

    def __iter__(self):
        return self

    def __next__(self):
        """ Return dict as required """
        row = next(self.rows)
        row_content = {}
        for cell in row:
            row_content[self.mapping[cell.column]] = cell.value
        # logger.debug("Data: %s", row_content)
        return row_content
